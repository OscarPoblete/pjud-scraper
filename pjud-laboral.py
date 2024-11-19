import requests
from bs4 import BeautifulSoup
import pandas as pd
import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Imports para manejar caracteres ilegales en los datos
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

url = "https://www.pjud.cl/ajax/Courts/getDailyStatements"

with open("TRIBUNALES_laboral.json", "r", encoding="utf-8") as file: #lee el archivo TRIBUNALES.json generado por pjud-tribunales.py
    tribunales = json.load(file)

fechas = ["29-09-2024", "30-09-2024", "01-10-2024", "02-10-2024", "03-10-2024", "04-10-2024", "05-10-2024", "06-10-2024", "07-10-2024", "08-10-2024", "09-10-2024", "10-10-2024", "11-10-2024", "12-10-2024", "13-10-2024", "14-10-2024", "15-10-2024", "16-10-2024", "17-10-2024"]

try:
    wb = load_workbook("Estados_diarios_LABORAL.xlsx")
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    #ws.append(["N°", "Número de Ingreso", "Partes", "Providencias", "TRIBUNAL", "Fecha"])


for tribunal in tribunales:
    data = []
    print(f"*Procesando datos de {tribunal['nombre_tribunal']}...")
    for fecha in fechas:

        payload = {
            "cod_tribunal": tribunal["cod_tribunal"],
            "tipo_juzgado": tribunal["tipo_juzgado"],
            "nombre_tribunal": tribunal["nombre_tribunal"],
            "date": fecha
        }

        # Making the POST request
        response = requests.post(url, data=payload)
        response.raise_for_status()  # Check for request errors

        # Parse the HTML content
        soup = BeautifulSoup(response.text, "html.parser")

        table_laboral = soup.find("table", id="data-table-estado-diario-laboral")
        table = soup.find("table", id="data-table-estado-diario") #cuando hay una unica tabla en el tribunal
        
        if table_laboral:
            print(f"  - Fecha: {fecha} - OK")
            # Extract table headers
            headers = [header.text.strip() for header in table_laboral.find_all("th")]

            # Extract table rows
            for row in table_laboral.find_all("tr")[1:]:  # Skip header row
                columns = [col.text.strip() for col in row.find_all("td")]
                data.append(columns + [tribunal["nombre_tribunal"], fecha])

        elif table:
            print(f"  - Fecha: {fecha} - OK (data-table-estado-diario)")
            # Extract table headers
            headers = [header.text.strip() for header in table.find_all("th")]

            # Extract table rows
            for row in table.find_all("tr")[1:]:  # Skip header row
                columns = [col.text.strip() for col in row.find_all("td")]
                data.append(columns + [tribunal["nombre_tribunal"], fecha])
        else:
            print(f"  - Fecha: {fecha} - no hay datos")
        
    if data:
        headers.extend(["TRIBUNAL", "Fecha"])
        df = pd.DataFrame(data, columns=headers)
        for row in dataframe_to_rows(df, index=False, header=False):
            try:
                ws.append(row)
            except IllegalCharacterError: #si el dato viene con algun caracter ilegal, lo reemplaza por un espacio
                with open("caracteres_ilegales.txt", "a", encoding="utf-8") as file:
                    file.write(f"Tribunal: {tribunal["nombre_tribunal"]}, {row}\n")
                cleaned_row = [ILLEGAL_CHARACTERS_RE.sub("", str(cell)) for cell in row]
                ws.append(cleaned_row)
        wb.save("Estados_diarios_LABORAL.xlsx")
        print(f"Estados diarios de '{tribunal['nombre_tribunal']}' guardados")
    else:
        print(f"No hay estados diarios para {tribunal['nombre_tribunal']}.")

print("Terminado.")